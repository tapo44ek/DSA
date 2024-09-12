import certifi
import requests
from aiogram import Bot
from aiogram.enums import ParseMode
from bs4 import BeautifulSoup
import pandas as pd
import time
from datetime import datetime
from multiprocessing import Process, Queue
from datetime import datetime, timedelta
import json

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import numpy as np
import warnings
import selenium
import shutil
import os, os.path
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
import datetime
import re
from datetime import datetime, timedelta
# import win32com.client as win32
import json
from selenium.webdriver.common.action_chains import ActionChains
import SMTPmail
import zipfile
import data_module
# from handlers import bot
import config
import asyncio
from aiogram.types import Message, FSInputFile
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import aiogram
bot = Bot(token=config.BOT_TOKEN, parse_mode=ParseMode.HTML)


def make_file(path, filename):
    df_M = pd.read_excel(os.path.join(path, 'М.xlsx'))
    df_M['№, дата регистрации'] = df_M['№, дата регистрации'].apply(lambda x: x.split(' ')[0])
    df_B = pd.read_excel(os.path.join(path, 'Б.xlsx'))
    df_B['№, дата регистрации'] = df_B['№, дата регистрации'].apply(lambda x: x.split(' ')[0])
    df_Shtatka = pd.read_excel(os.path.join(path, '2023.01.17 Штатка по блоку БРГ.xlsx'))
    print(df_M.columns)
    df_final = df_B
    # df_final['Срок ОА'] = None
    df_final.rename(columns={'Срок исполнения': 'Срок РГ'}, inplace=True)
    df_M.rename(columns={'Срок исполнения': 'Срок ОА'}, inplace=True)
    print(df_M.columns)
    df_final = df_final.merge(df_M[['№, дата регистрации', 'Срок ОА', 'Конечный исполнитель ветки резолюций']],
                              how='left', on='№, дата регистрации')
    # df_final['№, дата регистрации'] = df_final['№, дата регистрации'].apply(lambda x: x.split(' ')[0])
    print(df_final.head(5))
    list_delete = df_B['№, дата регистрации'].to_list()
    df_M.rename(columns={'Конечный исполнитель ветки резолюций': 'Конечный исполнитель ветки резолюций_x'},
                inplace=True)
    df_final = pd.concat([df_final, df_M[~df_M['№, дата регистрации'].isin(list_delete)]], axis=0)

    # df_final.drop_duplicates(subset=['', ''], inplace=True)
    # Функция для замены значений
    df_final['Конечный исполнитель ветки резолюций_x'] = np.where(
        ~df_final['Конечный исполнитель ветки резолюций_y'].isna(),
        df_final['Конечный исполнитель ветки резолюций_y'], df_final['Конечный исполнитель ветки резолюций_x'])
    # df_final['Конечный исполнитель ветки резолюций_x'] = df_final.apply(lambda row: row['Конечный исполнитель ветки резолюций_y'] if row['Конечный исполнитель ветки резолюций_y'] is not None else row['Конечный исполнитель ветки резолюций_x'], axis=1)

    df_final = df_final.merge(df_Shtatka[['ФИО короткий', 'Руководитель Зам']],
                              how='left', left_on='Конечный исполнитель ветки резолюций_x', right_on='ФИО короткий')

    list_workers = ['Мусиенко Ольга Александровна', 'Силуянова Юлия Павловна',
                    'Гибадулин Марат Мансурович', 'Лихач Ирина Андреевна']

    df_final = df_final[df_final['Руководитель Зам'].isin(list_workers)]
    df_final.drop(columns=['Конечный исполнитель ветки резолюций_y', 'ФИО короткий'], inplace=True)
    df_final['Срок РГ'] = pd.to_datetime(df_final['Срок РГ'], dayfirst=True)
    df_final['Срок ОА'] = pd.to_datetime(df_final['Срок ОА'], dayfirst=True)
    print(df_final.columns)
    df_final.rename(columns={'Конечный исполнитель ветки резолюций_x': 'Исполнитель блока',
                             'Руководитель Зам': 'Руководитель сотрудника'}, inplace=True)
    new_order = ['№ п/п', '№, дата регистрации', 'Откуда поступило, автор обращения',
                 '№, дата документа', 'Краткое содержание, что поручено', 'Куратор',
                 'Срок РГ', 'Срок ОА', 'Исполнитель блока', 'Руководитель сотрудника']
    df_final = df_final[['№ п/п', '№, дата регистрации', 'Откуда поступило, автор обращения',
                         '№, дата документа', 'Краткое содержание, что поручено', 'Куратор',
                         'Срок РГ', 'Срок ОА', 'Исполнитель блока', 'Руководитель сотрудника']]
    df_final.sort_values(by=['Срок РГ', 'Срок ОА'], inplace=True)
    df_final.drop_duplicates(subset=['№, дата регистрации', 'Исполнитель блока'], inplace=True)
    df_final['№ п/п'] = range(1, len(df_final) + 1)

    output_path = os.path.join(path, str(filename))
    df_final.to_excel(output_path, index=False)

    # Открываем файл Excel с помощью openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Установить ширину столбцов
    column_widths = {
        'A': 6,
        'B': 16,
        'C': 16,
        'D': 16,
        'E': 40,
        'F': 16,  # Ширина для столбца 'Name'
        'G': 18,  # Ширина для столбца 'Date'
        'H': 18,
        'I': 18,
        'J': 25,  # Ширина для столбца 'Sales'
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Создание стиля границ
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Применение границ ко всем ячейкам
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Сохранение изменений
    wb.save(output_path)

    print("Файл успешно сохранён с заданной шириной столбцов и границами.")
    return output_path


def sogly(s, DNSID, page, queue):
    url = f'https://mosedo.mos.ru/document.php?perform_search=1&DNSID={DNSID}&page={page}'
    r1 = s.get(url, verify=certifi.where())
    soup = BeautifulSoup(r1.text, 'html.parser')
    table = soup.find_all('table')
    df_foo = pd.read_html(str(table))[0]
    queue.put(df_foo)


def move_xl(ispolnit, finpath, macroname, last_msg_id):
    if macroname == 'Контроль общий.xlsm':
        ispolnit = str(ispolnit[0])
    path = os.path.join(os.getcwd(), 'downloads', 'mail_control')
    tech_path = os.path.join(workdir, 'income data', 'technical_data')
    tech_path = tech_path + '\\'
    files = os.listdir(path)
    files = [os.path.join(path, file) for file in files]
    files = [file for file in files if os.path.isfile(file)]
    fp = max(files, key=os.path.getctime)
    print(fp)
    (df_in,) = pd.read_html(fp)
    try:
        os.remove(fp)
    except:
        pass
    print(df_in)
    df_in.to_excel(os.path.join(finpath, ispolnit + ".xlsx"), index=False)
    # bot.send_message(UserID, str(fp))
    # try:
    #     os.rename(fp, tech_path + "02.xls")
    # except:
    #     print('ERROR, files to delete')
    #     # bot.send_message(UserID, 'Перезаписываю файлы ' + str(ispolnit))
    #     os.remove(tech_path + "02.xls")
    #     os.rename(fp, tech_path + "02.xls")
    # fname = tech_path + "02.xls"
    # # bot.send_message(UserID, 'xls -> xlsx ' + str(ispolnit))
    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # wb = excel.Workbooks.Open(fname)
    # try:
    #     os.remove(tech_path + "02.xlsx")
    # except:
    #     print('VSE HOROSHO')
    #
    # wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    # wb.Close()  # FileFormat = 56 is for .xls extension
    # os.remove(tech_path + "02.xls")
    # excel.Application.Quit()
    # xl = win32.Dispatch("Excel.Application")
    # xl.Workbooks.Open(os.path.abspath(tech_path + macroname), ReadOnly=1)
    # xl.Application.Quit()  # Comment this out if your excel script closes
    # del xl
    # try:
    #     os.rename(tech_path + '02.xlsx', finpath + ispolnit + ".xlsx")
    # except:
    #     print('ERROR, files to delete')
    #     # bot.send_message(UserID, 'Перезаписываю файлы '+ str(ispolnit))
    #     os.remove(finpath + ispolnit + ".xlsx")
    #     os.rename(tech_path + '02.xlsx', finpath + ispolnit + ".xlsx")
    # print('MACRO is DONE')
    last_msg_id = asyncio.run(info_msg(UserID, 'xlsx готов для ' + str(ispolnit), last_msg_id))
    # bot.send_message(UserID, 'xlsx готов для ' + str(ispolnit))
    return last_msg_id


def download_xls(ispolnit, date1, date2, finpath, macroname, PCuser, last_msg_id):
    driver.find_element(By.XPATH, '//*[@id="s-menu-stat"]').click()
    driver.find_element(By.XPATH, '//*[@id="s-menu-stat"]/div/a[1]').click()
    time.sleep(1)
    driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/a[36]').click()
    time.sleep(2)
    startdate = driver.find_element(By.XPATH,
                                    '//*[@id="fDetailedControl"]/div/div[1]/div[1]/div[1]/div[1]/div[1]/div/input')
    enddate = driver.find_element(By.XPATH,
                                  '//*[@id="fDetailedControl"]/div/div[1]/div[1]/div[1]/div[2]/div[1]/div/input')
    startdate.click()
    # startdate.send_keys(Keys.COMMAND, 'a')
    startdate.send_keys(Keys.CONTROL, 'a')
    startdate.send_keys(Keys.BACKSPACE)
    startdate.send_keys(str(date1))
    enddate.click()
    # enddate.send_keys(Keys.COMMAND, 'a')
    enddate.send_keys(Keys.CONTROL, 'a')
    enddate.send_keys(Keys.BACKSPACE)
    enddate.send_keys(str(date2))
    ispolnitel = driver.find_element(By.XPATH,
                                     '//*[@id="fDetailedControl"]/div/div[1]/div[2]/div/div/div[1]/div/div[2]/input')
    # ispolnitel_spisok = driver.find_element(By.XPATH, '//*[@id="fDetailedControl"]/div/div[1]/div[2]/div/div[1]/ul/li[2]')
    startbutton = driver.find_element(By.XPATH, '//*[@id="fDetailedControl"]/div/div[3]/button[1]')
    time.sleep(5)
    isp = ispolnit.split(sep='.')
    ispo = str(isp[0]).split()
    ispolnitel.click()
    ispolnitel.send_keys(str(ispo[0]))
    time.sleep(0.5)
    ispolnitel.send_keys(' ')
    time.sleep(0.5)
    ispolnitel.send_keys(str(ispo[1]))
    time.sleep(0.5)
    ispolnitel.send_keys('.')
    time.sleep(0.5)
    ispolnitel.send_keys(str(isp[1]) + '.')
    time.sleep(0.5)
    ispolnitel_spisok = driver.find_element(By.XPATH,
                                            '//*[@id="fDetailedControl"]/div/div[1]/div[2]/div/div/div[1]/ul/li[1]')
    g = 0
    while g != 10:
        if str(driver.find_element(By.XPATH,
                                   '//*[@id="fDetailedControl"]/div/div[1]/div[2]/div/div/div[1]/ul/li[1]/div/span[2]').text) == ispolnit:
            ispolnitel_spisok.click()
            g = 10
        else:
            # ispolnitel.send_keys(Keys.COMMAND, 'a')
            ispolnitel.send_keys(Keys.CONTROL, 'a')
            ispolnitel.send_keys(Keys.BACKSPACE)
            ispolnitel.send_keys(str(ispo[0]))
            time.sleep(0.5)
            ispolnitel.send_keys(' ')
            time.sleep(0.5)
            ispolnitel.send_keys(str(ispo[1]))
            time.sleep(0.5)
            ispolnitel.send_keys('.')
            time.sleep(0.5)
            ispolnitel.send_keys(str(isp[1]) + '.')
            time.sleep(2)
            # time.sleep(1)
    last_msg_id = asyncio.run(info_msg(UserID, "Загружаю страницу 1 из 2 для " + str(ispolnit), last_msg_id))
    # bot.send_message(UserID, "Загружаю страницу 1 из 2 для " + str(ispolnit))
    startbutton.click()
    try:
        WebDriverWait(driver, 600).until(EC.visibility_of_element_located(
            (By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/div/div[3]/div[2]/table/tbody')))
    except:
        print("Timed out waiting for page to load")
        last_msg_id = asyncio.run(info_msg(UserID, "Timeisout", last_msg_id))
        # bot.send_message(UserID, "Timeisout")
    finally:
        print("Page loaded")
        last_msg_id = asyncio.run(info_msg(UserID, "Страница 1 из 2 загружена", last_msg_id))
        # bot.send_message(UserID, "Страница 1 из 2 загружена")
    time.sleep(2)
    maxrow = len(driver.find_elements(By.XPATH,
                                      '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/div/div[3]/div[2]/table/tbody/tr'))
    maxcol = len(driver.find_elements(By.XPATH,
                                      '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/div/div[3]/div[2]/table/tbody/tr[' + str(
                                          maxrow) + ']/td'))
    print(maxrow)
    print(maxcol)
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(3)
    last_msg_id = asyncio.run(info_msg(UserID, "Загружаю страницу 2 из 2 для " + str(ispolnit), last_msg_id))
    # bot.send_message(UserID, "Загружаю страницу 2 из 2 для " + str(ispolnit))
    driver.find_element(By.XPATH,
                        '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/div/div[3]/div[2]/table/tbody/tr[' + str(
                            maxrow) + ']/td[' + str(maxcol) + ']/a').click()
    time.sleep(4)
    try:
        WebDriverWait(driver, 600).until(
            EC.visibility_of_element_located((By.XPATH, '//a[contains(@href,"to=excel")]')))
    except:
        print("Timed out waiting for page to load")
        last_msg_id = asyncio.run(info_msg(UserID, "Timeisout" + str(ispolnit), last_msg_id))
        # bot.send_message(UserID, "Timeisout")
    finally:
        last_msg_id = asyncio.run(info_msg(UserID, "Страница 2 из 2 загружена" + str(ispolnit), last_msg_id))
        # bot.send_message(UserID, "Страница 2 из 2 загружена")
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)
    path1 = os.path.join(os.getcwd(), 'downloads', 'mail_control')
    files = os.listdir(path1)
    files = [os.path.join(path1, file) for file in files]
    files = [file for file in files if os.path.isfile(file)]
    filename = max(files, key=os.path.getctime)
    print(filename)
    driver.find_element(By.XPATH, '//a[contains(@href,"to=excel")]').click()
    #######################################################
    files = os.listdir(path1)
    files = [os.path.join(path1, file) for file in files]
    files = [file for file in files if os.path.isfile(file)]
    filename1 = max(files, key=os.path.getctime)
    print(filename1)
    while filename1 == filename:
        time.sleep(10)
        files = os.listdir(path1)
        files = [os.path.join(path1, file) for file in files]
        files = [file for file in files if os.path.isfile(file)]
        filename1 = max(files, key=os.path.getctime)
        print(filename1)
    time.sleep(10)
    last_msg_id = move_xl(ispolnit, finpath, macroname, last_msg_id)

    return last_msg_id


def copy_net(folder_from, folder_to):
    for f in os.listdir(folder_from):
        if os.path.isfile(os.path.join(folder_from, f)):
            try:
                shutil.copy(os.path.join(folder_from, f), os.path.join(folder_to, f))
            except:
                print('Такой файл уже есть, копирую с заменой')
                os.remove(os.path.join(folder_to, f))
                shutil.copy(os.path.join(folder_from, f), os.path.join(folder_to, f))
            if os.path.isdir(os.path.join(folder_from, f)):
                os.system(f'rd /S /Q {folder_to}\\{f}')
                try:
                    shutil.copytree(os.path.join(folder_from, f), os.path.join(folder_to, f))
                except:
                    print('Такой файл уже есть, копирую с заменой')
                    os.remove(os.path.join(folder_to, f))
                    shutil.copytree(os.path.join(folder_from, f), os.path.join(folder_to, f))

        # exec(open("C:\\Users\\ArsenevVD\\Desktop\\JupiterLab\\selenium\\part1.py").read())
        # os.system('python C:\\Users\\ArsenevVD\\Desktop\\JupiterLab\\selenium\\part1.py' )


async def info_msg(tg_id, text, last_msg_id):
    if last_msg_id != 0:
        await bot.delete_message(chat_id=tg_id, message_id=last_msg_id)
        try:
            await bot.session.close()
        except:
            pass
    last_msg_id = await bot.send_message(chat_id=tg_id, text=text)
    last_msg_id = last_msg_id.message_id
    try:
        await bot.session.close()
    except:
        pass
    return last_msg_id

if __name__ == '__main__':
    print(os.getcwd())
    last_msg_id = 0
    warnings.filterwarnings('ignore')
    ds = datetime.now()
    print(ds)
    with open(os.path.join(os.getcwd(), 'settings.json')) as f:
        settings = json.load(f)
        print(settings)
        token = settings['token2']
        SEDOlog = settings['SEDOlog']
        SEDOpass = settings['SEDOpass']
        PCuser = settings['PCuser']
        UserID = settings['UserID']
        folder_to = settings['mailFolder']
        folder_control = settings['mail_folder_final']
    UserID = data_module.get_report('control_mail')
    print(type(UserID))
    print('TG_ID is: ', UserID)
    workdir = os.getcwd()
    last_msg_id = asyncio.run(info_msg(UserID, 'Начинаю контроль писем', last_msg_id))
    # last_msg_id = asyncio.run(bot.send_message(chat_id=UserID, text='Начинаю контроль писем')).message_id
    # try:
    #     asyncio.run(bot.session.close())
    # except:
    #     pass

    now = datetime.today()
    now.date
    end_date = now + timedelta(days=30)
    date2 = str(end_date.date())
    date2 = date2.split(sep='-')
    date1 = str(now.date())
    date1 = date1.split(sep='-')
    date22 = date2[2] + '.' + date2[1] + '.' + date2[0]
    date11 = date1[2] + '.' + date1[1] + '.' + date1[0]
    date222 = date2[1] + '.' + date2[2]
    date111 = date1[1] + '.' + date1[2]
    startdateM = now - timedelta(days=90)
    date3 = str(startdateM.date())
    date3 = date3.split(sep='-')
    date33 = date3[2] + '.' + date3[1] + '.' + date3[0]
    if os.path.exists(os.path.join(os.getcwd(), 'export_data', 'Контроль писем общий')) == False:
        os.makedirs(os.path.join(os.getcwd(), 'export_data', 'Контроль писем общий'))
        # os.rename()
    finpath = os.path.join(os.getcwd(), 'export_data', 'Контроль писем общий')
    foldername = os.path.join(finpath, "Контроль " + str(date111) + '-' + str(date222))

    if os.path.exists(foldername) == False:
        os.mkdir(foldername)
        shutil.copy(os.path.join(workdir, 'income data', 'technical_data', 'Контроль.xlsx'),
                    foldername)
        shutil.copy(os.path.join(workdir, 'income data', 'technical_data', 'Список всех сотрудников УВУЖ.xlsx'),
                    foldername)
        shutil.copy(os.path.join(workdir, 'income data', 'technical_data', 'Макрос.xlsm'),
                    foldername)
        shutil.copy(os.path.join(workdir, 'income data', 'technical_data', '2023.01.17 Штатка по блоку БРГ.xlsx'),
                    foldername)
    folder_from = os.path.join(os.getcwd(), 'export_data', 'Контроль писем общий')


    incomepath = os.path.join(workdir, 'income data', 'income_data_D.xlsx')
    incomedf = pd.read_excel(incomepath, 'sheet1')
    User = str(incomedf.iloc[0, 5])
    ispolniteli = incomedf['Фамилия И.О. контроль общий'].tolist()
    print(ispolniteli)
    ispolniteli = [item for item in ispolniteli if not (pd.isnull(item)) == True]
    print(ispolniteli)

    path = os.getcwd()
    chrome_options = webdriver.ChromeOptions()
    prefs = {'profile.default_content_settings.popups': 0,
             "download.default_directory": os.path.join(os.getcwd(), 'downloads', 'mail_control'),
             "directory_upgrade": True}
    chrome_options.add_experimental_option('prefs', prefs)

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.get('https://mosedo.mos.ru/auth.php?uri=%2F')
    time.sleep(3)
    organization = driver.find_element(By.XPATH, '//*[@id="organizations"]')
    user = driver.find_element(By.XPATH, '//*[@id="logins"]')
    psw = driver.find_element(By.XPATH, '//*[@id="password_input"]')
    login = driver.find_element(By.XPATH, '//*[@id="login_form"]/table/tbody/tr[11]/td[4]/input[1]')
    organization.send_keys('Департамент городского имущества города Москвы')
    time.sleep(2)
    organization1 = driver.find_element(By.XPATH,
                                        '//*[@id="ui-id-1"]/li/a[text() = "Департамент городского имущества города Москвы"]')
    organization1.click()
    time.sleep(3)
    user.send_keys(SEDOlog)
    time.sleep(5)
    user1 = driver.find_element(By.XPATH, '//*[@id="ui-id-2"]/li/a[text() = "' + str(SEDOlog) + '"]')
    user1.click()
    time.sleep(3)
    psw.send_keys(SEDOpass)
    login.click()
    time.sleep(2)
    print(ispolniteli[0])

    for i in range(len(ispolniteli)):
        isp = str(ispolniteli[i])
        # if isp != 'Биктимиров Р.Г.':
        macroname = 'Контроль общий.xlsm'
        last_msg_id = download_xls(isp, date11, date22, foldername, macroname, PCuser, last_msg_id)

    year1 = 2020
    year2 = now.year
    d_start = '01.01.2019'
    date_now = datetime.strftime(datetime.now(), '%d.%m.%Y')
    # chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("--headless")

    # driver = webdriver.Chrome(options=chrome_options)
    # driver = webdriver.Chrome()

    # url_kontrol = f'https://mosedo.mos.ru/auth.php?uri=%2Fstat%2Fcontrol_stats.details.php%3Ffixed%3D%26delegate_id%3D%26is_letter%3D%26report_name%3Dcontrol_stats%26ctl_type%255B0%255D%3D0%26ctl_type%255B1%255D%3D1%26later_type%3D0%26due_date_from%3D{d_start}%26due_date_until%3D{d_end}%26start_rdate%3D%26end_rdate%3D%26user%255B0%255D%3D0%26inv_user%255B0%255D%3D0%26executor%3D{EXECUTOR_ID}%26inv_executor%3D0%26result%3D%25D1%25F4%25EE%25F0%25EC%25E8%25F0%25EE%25E2%25E0%25F2%25FC%2B%25EE%25F2%25F7%25E5%25F2...'
    url_auth = 'https://mosedo.mos.ru/auth.php?group_id=21'

    s = requests.Session()

    DNSID = 'wJBIu0NhWLwF8YslRiym5mw'
    url_control = f'https://mosedo.mos.ru/stat/reports_msk.php?DNSID={DNSID}'


    url_sogl = f'https://mosedo.mos.ru/document.php?all=1&category=6&DNSID={DNSID}&whole_period=1&ajax=1&page=1&isJson=0&DNSID={DNSID}'

    DNSID = 'wJBIu0NhWLwF8YslRiym5mw'

    data = {"DNSID": DNSID,
            "group_id": "21",
            "login": "%C0%F0%F1%E5%ED%FC%E5%E2+%C2.%C4.",
            "user_id": "80742170",
            "password": SEDOpass,  # ПОМЕНЯТЬ НА ДАННЫЕ ИЗ ФАЙЛА КОНФИГУРАЦИИ
            "token": "",
            "x": "1"}

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:109.0) Gecko/20100101 Firefox/114.0',
        'Connection': 'keep-alive'}

    data_DS = {
        'check_all_documents': 'on',
        'type_0': '1',
        'type_1': '1',
        'type_2': '1',
        'type_3': '1',
        'type_12': '1',
        'type_13': '1',
        'type_4': '1',
        'type_5': '1',
        'has_period': '1',
        'year_from': '2020',
        'year_to': str(year2),
        'org_name': '%C4%C3%C8%E3%CC',
        'org': '21',
        'order_by': 'default',
        'required_text': '',
        'num': '',
        'rdate_f': '',
        'reg_user': '',
        'reg_user_id': '',
        'rdate_t': '',
        'recipient': '',
        'recipient_id': '',
        'recipient_group': '',
        'recipient_group_id': '',
        'in_number': '',
        'bound_number': '',
        'contract_bound_number': '',
        'recipient_org_id': '',
        'cl_out_num': '',
        'cl_out_date_f': '',
        'cl_out_date_t': '',
        'cl_sign': '',
        'cl_sign_id': '',
        'cl_sign_group': '',
        'cl_sign_group_id': '',
        'cl_executor': '',
        'cl_executor_id': '',
        'cl_executor_group': '',
        'cl_executor_group_id': '',
        'cl_text': '',
        'out_number': '',
        'out_date_f': '',
        'out_reg_user': '',
        'out_reg_user_id': '',
        'out_date_t': '',
        'author': '',
        'author_id': '',
        'author_group': '',
        'author_group_id': '',
        'prepared_by': '',
        'prepared_by_id': '',
        'prepared_by_org_id': '',
        'curator': '',
        'curator_id': '',
        'short_content': '',
        'document_kind': '0',
        'delivery_type': '',
        'document_special_kind': '0',
        'external_id': '',
        'has_manual_sign': '0',
        'is_hand_shipping': '0',
        'sign_type': '0',
        'is_dsp': '0',
        'is_control': '0',
        'is_urgent': '0',
        'creator': '',
        'creator_id': '',
        'memo': '',
        'send_date_f': '',
        'send_date_t': '',
        'info': '',
        'info_author': '',
        'info_author_id': '',
        'info_date_f': '',
        'info_date_t': '',
        'og_file_number': '0',
        'rec_vdelo': '0',
        'vdelo_date_f': '',
        'vdelo_date_t': '',
        'vdelo_prepared': '',
        'vdelo_prepared_id': '',
        'vdelo_signed': '',
        'vdelo_signed_id': '',
        'vdelo_text': '',
        'res_type': '0',
        'res_urgency': '0',
        'resolution_num': '',
        'r_rdate_f': str(date33),  ############# Date of document start
        'resolution_creator': '',
        'resolution_creator_id': '',
        'r_rdate_t': '',
        'resolution_author': '',
        'resolution_author_id': '',
        'resolution_author_group': '',
        'resolution_author_group_id': '',
        'resolution_author_org_id': '',
        'r_special_control': '0',
        'resolution_behalf': '',
        'resolution_behalf_id': '',
        'resolution_acting_author': '',
        'resolution_acting_author_id': '',
        'resolution_to': '%CC%F3%F1%E8%E5%ED%EA%EE+%CE.%C0.',
        'resolution_to_id': '70045',
        'resolution_to_group': '%C4%E5%EF%E0%F0%F2%E0%EC%E5%ED%F2+%E3%EE%F0%EE%E4%F1%EA%EE%E3%EE+%E8%EC%F3%F9%E5%F1%F2%E2%E0+%E3%EE%F0%EE%E4%E0+%CC%EE%F1%EA%E2%FB',
        'resolution_to_group_id': '21',
        'resolution_to_org_id': '',
        'res_project_letter': '0',
        'res_curator': '',
        'res_curator_id': '',
        'r_control': '0',
        'r_control_f': '',
        'r_control_t': '',
        'r_otv': '0',
        'r_dback': '0',
        'resolution_text': '',
        'r_ef_reason_category_id': '0',
        'r_ef_reason_id': '0',
        'r_is_signed': '0',
        'r_plus': '0',
        'r_another_control': '0',
        'r_oncontrol': '0',
        'r_oncontrol_f': '',
        'r_oncontrol_t': '',
        'unset_control': '0',
        'unset_control_f': '',
        'unset_control_t': '',
        're_date_f': '',
        're_date_t': '',
        're_author': '',
        're_author_id': '',
        're_author_group': '',
        're_author_group_id': '',
        're_acting_author': '',
        're_acting_author_id': '',
        're_is_interim': '-1',
        're_text': '',
        'docs_in_execution': '0',
        're_doc_org_id': '',
        'csdr_initiator': '',
        'csdr_initiator_id': '',
        'csdr_initiator_group': '',
        'csdr_initiator_group_id': '',
        'csdr_start': '0',
        'csdr_stop': '0',
        'and[csdr][0]': '0',
        'participant_name_0': '',
        'participant_name_0_id': '',
        'participant_group_0': '',
        'participant_group_0_id': '',
        'csdr_has_deadline_0': '0',
        'csdr_status_0': '0',
        'csdr_init_date_0_f': '',
        'csdr_init_date_0_t': ''
    }


    print('1')
    r = s.post('https://mosedo.mos.ru/auth.php?group_id=21', data=data, headers=headers, allow_redirects=False)
    DNSID = r.headers['location'].split('DNSID=')[1]
    # auth_token = s.cookies.get_dict()['auth_token']
    url_sogl = f'https://mosedo.mos.ru/document_search.php?new=0&DNSID={DNSID}'
    
    headers2 = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:109.0) Gecko/20100101 Firefox/115.0',
        'Connection': 'keep-alive',
        'Sec-Fetch-User': '?1',
        'Sec-Fetch-Dest': 'document'
    }
    headers2 = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:109.0) Gecko/20100101 Firefox/115.0',
        'Connection': 'keep-alive',
        'Sec-Fetch-User': '?1',
        'Sec-Fetch-Dest': 'document'
    }
    print(r)
    s.cookies
    driver.close()
    r2 = s.post(url_sogl, data=data_DS, headers=headers2)
    print('2')
    first_soup = BeautifulSoup(r2.text, 'html.parser')

    # with open("sogl.html", "w") as file:
    #     file.write(r2.text)
    table = first_soup.find_all('table')
    count_doc = int(first_soup.find('span', attrs={'class': "search-export__count"}).text.split()[-1])
    count_page = count_doc // 15 + 1
    print(f'{count_doc} equal {count_page} pages')
    df = pd.read_html(str(table))[0]
    df.to_excel('table.xlsx')

    i = 1
    k = 0

    queue = Queue()
    processes = []
    df_list = []
    df_final = pd.DataFrame()
    count_cycle = count_page // 50 + 1
    for k in range(count_cycle):
        for i in range(k * 50 + 1, (k + 1) * 50 + 1):
            p = Process(target=sogly, args=(s, DNSID, i, queue))
            p.start()
            processes.append(p)

        for i in range(k * 50 + 1, (k + 1) * 50 + 1):
            df = queue.get()
            df_list.append(df)

        for p in processes:
            p.join()

    df_final = pd.concat(df_list)
    df_final = df_final.drop(columns=['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2'])
    df_final.dropna(subset=['№ Дата рег.'], inplace=True)

    ########################################################################

    df1 = df_final
    df1 = df1.drop(columns=['Тип'])
    df1['Адресат Сопр. письмо, поручение Автор'] = df1['Адресат Сопр. письмо, поручение Автор'].str.replace('Кому: ',
                                                                                                            '')
    new_df_1 = df1['№ Дата рег.'].str.split('  ', expand=True)
    new_df_1 = new_df_1.rename(columns={0: '№ документа', 1: 'Дата регистрации'})

    new_df_2 = df1['Адресат Сопр. письмо, поручение Автор'].str.split('От кого: ', expand=True)
    new_df_2 = new_df_2.rename(columns={0: 'Адресат', 1: 'Автор'})
    new_df_2.info()
    df1 = pd.concat([df1, new_df_1, new_df_2], axis=1)
    df1.drop(columns=['Адресат Сопр. письмо, поручение Автор', '№ Дата рег.'], inplace=True)
    df1 = df1[['№ документа', 'Дата регистрации', 'Исх № Дата рег.', 'Адресат', 'Автор', 'Краткое содержание']]

    df1 = df1.set_index(np.arange(1, len(df1) + 1))
    df1.index.rename('№ п/п', inplace=True)

    df1.to_excel(os.path.join(foldername, 'Мусиенко О.А..xlsx'))
    df1.info()
    de = datetime.now()
    print(de)
    print('\n')
    print(de - ds)
    f = "Контроль " + str(date111) + '-' + str(date222)
    filename = f'Контроль {date111}-{date222}.xlsx'

    att = make_file(os.path.join(os.getcwd(), 'export_data', 'Контроль писем общий', f), filename)


    files = os.listdir(os.path.join(folder_from, f))
    archive = 'mail_control.zip'
    with zipfile.ZipFile(archive, "w") as zf:
        for file in files:
            file_path = os.path.join(folder_from, f, file)
            zf.write(file_path, arcname=os.path.relpath(file_path, folder_from))
    last_msg_id = asyncio.run(info_msg(UserID, "Выгрузка завершена", last_msg_id))
    att_paths = ['mail_control.zip', att]
    for i in range(len(att_paths)):
        print(att_paths[i])
        file = FSInputFile(att_paths[i])
        asyncio.run(bot.send_document(UserID, file, caption=f'Файлы для контроля писем'))
        try:
            asyncio.run(bot.session.close())
        except:
            pass
    # login = r"HQ\ArsenevVD"
    # password = "Vitosik02010201!"
    # server = "owa.mos.ru"
    # port = 587
    # recipients = ['ArsenevVD@mos.ru', 'GoncharovaIA2@mos.ru', 'GabitovDS@mos.ru']
    # cc = []
    # subject = 'Файлы для контроля писем'
    # body = ''
    # attachment_paths = ['mail_control.zip']
    # SMTPmail.send_email(login, password, server, port, recipients, cc, subject, body, attachment_paths)
    # bot.send_message(UserID, 'Письмо направлено')
    # try:
    #     shutil.copytree(os.path.join(folder_from, f), os.path.join(folder_to, f))
    # except:
    #     print('папка уже существует либо есть проблема с сетевыми папками')
    #     bot.send_message(UserID, 'папка уже существует либо есть проблема с сетевыми папками')
    # bot.send_message(UserID, 'ALL IS DONE!!!!')
    # driver.close()
