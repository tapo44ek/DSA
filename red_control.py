# -*- coding: utf-8 -*-
import selenium
from selenium import webdriver
import shutil
import os, os.path
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime
import re
from datetime import datetime, timedelta
import win32com.client as win32
import pandas as pd
import numpy as np
import openpyxl as ox
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from openpyxl import load_workbook
import json
import SMTPmail
import data_module
from handlers import bot
import asyncio
from aiogram.types import Message, FSInputFile


async def info_msg(tg_id, text, last_msg_id):
    if last_msg_id != 0:
        await bot.delete_message(chat_id=tg_id, message_id=last_msg_id)
        try:
            await bot.session.close()
        except:
            pass
    last_msg_id = await bot.send_message(chat_id=tg_id, text=text)
    last_msg_id = last_msg_id.message_id
    try:
        await bot.session.close()
    except:
        pass
    return last_msg_id


# def move_xl(ispolnit, finpath, macroname, PCuser, last_msg_id):
#     if macroname == 'Контроль общий.xlsm':
#         ispolnit = str(ispolnit[0])
#     path = 'C:\\Users\\' + PCuser + '\\Downloads'
#     tech_path = os.path.join(workdir, 'income data', 'technical_data')
#     tech_path = tech_path + '\\'
#     techpath2 = os.path.join(workdir, 'income data', 'technical_data', 'red control')
#     techpath2 = techpath2 + '\\'
#     files = os.listdir(path)
#     files = [os.path.join(path, file) for file in files]
#     files = [file for file in files if os.path.isfile(file)]
#     fp = max(files, key=os.path.getctime)
#     print(fp)
#     # bot.send_message(UserID, str(fp))
#     try:
#         os.rename(fp, tech_path + "03.xls")
#     except:
#         print('ERROR, files to delete 1')
#         bot.send_message(UserID, 'Перезаписываю файлы ' + str(ispolnit))
#         os.remove(tech_path + "03.xls")
#         os.rename(fp, tech_path + "03.xls")
#     fname = tech_path + "03.xls"
#     # bot.send_message(UserID, 'xls -> xlsx ' + str(ispolnit))
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     wb = excel.Workbooks.Open(fname)
#     try:
#         os.remove(tech_path + "03.xlsx")
#     except:
#         print('VSE HOROSHO')
#         # bot.send_message(UserID, 'VSE HOROSHO')
#
#     wb.SaveAs(fname + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
#     wb.Close()  # FileFormat = 56 is for .xls extension
#     os.remove(tech_path + "03.xls")
#     excel.Application.Quit()
#     print('macrostart')
#     xl = win32.Dispatch("Excel.Application")
#     xl.Workbooks.Open(os.path.abspath(tech_path + macroname), ReadOnly=1)
#     xl.Application.Quit()  # Comment this out if your excel script closes
#     print('macroend')
#     del xl
#     try:
#         os.rename(tech_path + '03.xlsx', techpath2 + ispolnit + ".xlsx")
#     except:
#         print('ERROR, files to delete 2')
#         print(techpath2 + ispolnit + ".xlsx")
#         # bot.send_message(UserID, 'Перезаписываю файлы '+ str(ispolnit))
#         os.remove(techpath2 + ispolnit + ".xlsx")
#         os.rename(tech_path + '03.xlsx', techpath2 + ispolnit + ".xlsx")
#     print(finpath + ispolnit + ".xlsx")
#     print('MACRO is DONE')
#     last_msg_id = asyncio.run(info_msg(UserID, 'xlsx готов для ' + str(ispolnit), last_msg_id))
#     # bot.send_message(UserID, 'xlsx готов для ' + str(ispolnit))
#
#     return last_msg_id
#
#
# def move_xl1(ispolnit, finpath, macroname, PCuser, last_msg_id):
#     if macroname == 'Контроль общий.xlsm' :
#         ispolnit = str(ispolnit[0])
#     path = 'C:\\Users\\'+PCuser+'\\Downloads'
#     tech_path = os.path.join(workdir, 'income data', 'technical_data')
#     tech_path = tech_path + '\\'
#     techpath2 = os.path.join(workdir, 'income data', 'technical_data', 'pred control')
#     techpath2 = techpath2 + '\\'
#     files = os.listdir(path)
#     files = [os.path.join(path, file) for file in files]
#     files = [file for file in files if os.path.isfile(file)]
#     fp = max(files, key=os.path.getctime)
#     print (fp)
#     #bot.send_message(UserID, str(fp))
#     try:
#         os.rename(fp, tech_path + "03.xls")
#     except:
#         print ('ERROR, files to delete 1')
#         bot.send_message(UserID, 'Перезаписываю файлы '+ str(ispolnit))
#         os.remove(tech_path + "03.xls")
#         os.rename(fp, tech_path + "03.xls")
#     fname = tech_path + "03.xls"
#     #bot.send_message(UserID, 'xls -> xlsx ' + str(ispolnit))
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     wb = excel.Workbooks.Open(fname)
#     try:
#         os.remove(tech_path + "03.xlsx")
#     except:
#         print('VSE HOROSHO')
#         #bot.send_message(UserID, 'VSE HOROSHO')
#
#     wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
#     wb.Close()                               #FileFormat = 56 is for .xls extension
#     os.remove(tech_path + "03.xls")
#     excel.Application.Quit()
#     xl=win32.Dispatch("Excel.Application")
#     xl.Workbooks.Open(os.path.abspath(tech_path + macroname), ReadOnly=1)
#     xl.Application.Quit() # Comment this out if your excel script closes
#     del xl
#     try:
#         os.rename(tech_path + '03.xlsx', techpath2 + ispolnit + ".xlsx")
#     except:
#         print ('ERROR, files to delete 2')
#         print(techpath2 + ispolnit + ".xlsx")
#         #bot.send_message(UserID, 'Перезаписываю файлы '+ str(ispolnit))
#         os.remove(techpath2 + ispolnit + ".xlsx")
#         os.rename(tech_path + '03.xlsx', techpath2 + ispolnit + ".xlsx")
#     print(finpath + ispolnit + ".xlsx")
#     print('MACRO is DONE')
#     last_msg_id = asyncio.run(info_msg(UserID, 'xlsx готов для ' + str(ispolnit), last_msg_id))
#     # bot.send_message(UserID, 'xlsx готов для ' + str(ispolnit))
#     return last_msg_id


if __name__ == "__main__":
    last_msg_id = 0
    with open('settings.json') as f:
        settings = json.load(f)
        print(settings)
    #    time.sleep(20)
        token = settings['token2']
        SEDOlog = settings['SEDOlog']
        SEDOpass = settings['SEDOpass']
        PCuser = settings['PCuser']
        UserID = settings['UserID']
        folder_to = settings['mailFolder']
        folder_control = settings['mail_folder_final']
    UserID = data_module.get_report('red_control')
    print(type(UserID))
    print('TG_ID is: ', UserID)
    workdir = os.getcwd()
    incomepath = os.path.join(workdir, 'income data', 'income_data_D.xlsx')
    incomedf = pd.read_excel(incomepath, 'sheet1')
    User = str(incomedf.iloc[0,5])
    ispolniteli = incomedf['Фамилия И.О. контроль общий'].tolist()
    print(ispolniteli)
    ispolniteli = [item for item in ispolniteli if not(pd.isnull(item)) == True]

    # bot = telebot.TeleBot(token);
    # @bot.message_handler(content_types=['text'])
    # def get_text_messages(message):
    #     if message.text == "Привет":
    #         bot.send_message(message.from_user.id, "Привет, чем я могу тебе помочь?")
    #     elif message.text == "Тест":
    #         bot.send_message(message.from_user.id, "Запускаю скрипт")
    #         #exec(open("C:\\Users\\ArsenevVD\\Desktop\\JupiterLab\\selenium\\part1.py").read())
    #         #os.system('python C:\\Users\\ArsenevVD\\Desktop\\JupiterLab\\selenium\\part1.py' )
    # bot.send_message(UserID, '------------------------')
    # bot.send_message(UserID, '------------------------')
    # bot.send_message(UserID, 'Начинаю красный контроль')
    last_msg_id = asyncio.run(info_msg(UserID, 'Начинаю красный и предупредительный контроли', last_msg_id))
    now = datetime.today()
    now.date
    end_date = now - timedelta(days = 120)
    date2 = str(end_date.date())
    date2 = date2.split(sep='-')
    date1 = str(now.date())
    date1 = date1.split(sep='-')
    date22 = date2[2]+'.'+date2[1]+'.'+date2[0]
    date11 = date1[2]+'.'+date1[1]+'.'+date1[0]
    date222 = date2[1]+'.'+date2[2]
    date111 = date1[1]+'.'+date1[2]
    startdateM = now + timedelta(days = 30)
    date3 = str(startdateM.date())
    date3 = date3.split(sep='-')
    date33 = date3[2]+'.'+date3[1]+'.'+date3[0]
    date333 = date3[1]+'.'+date3[2]
    finpath = 'C:\\Users\\' + PCuser + '\\Desktop\\'

    if os.path.exists(finpath + 'Контроль писем общий') == False:
        os.mkdir(finpath + 'Контроль писем общий')
        #os.rename()
    finpath = 'C:\\Users\\' + PCuser + '\\Desktop\\Контроль писем общий\\'
    foldername = finpath + "Контроль " + str(date111) + '-' + str(date333) + '\\'
    if os.path.exists(foldername) == False:
        os.mkdir(foldername)
        # bot.send_message(UserID, 'Создал папку для красного контроля')
        # bot.send_message(UserID, foldername)




    path = os.getcwd()

    r = os.listdir(os.path.join(workdir, 'income data', 'technical_data', 'red control'))
    for i in range (len(r)):
        try:
            os.remove(str(os.path.join(workdir, 'income data', 'technical_data', 'red control'))+'\\'+str(r[i]))
        except:
            pass

    year1 = 2020
    year2 = now.year
    chromedriver = str(path) + '/chromedriver.exe'
    driver = webdriver.Chrome()
    driver.get('https://mosedo.mos.ru/auth.php?uri=%2F')
    time.sleep(3)
    organization = driver.find_element(By.XPATH, '//*[@id="organizations"]')
    user = driver.find_element(By.XPATH, '//*[@id="logins"]')
    psw = driver.find_element(By.XPATH, '//*[@id="password_input"]')
    login = driver.find_element(By.XPATH, '//*[@id="login_form"]/table/tbody/tr[11]/td[4]/input[1]')
    organization.send_keys('Департамент городского имущества города Москвы')
    time.sleep(2)
    organization1 = driver.find_element(By.XPATH, '//*[@id="ui-id-1"]/li/a[text() = "Департамент городского имущества города Москвы"]')
    organization1.click()
    user.send_keys(SEDOlog)
    time.sleep(3)
    user1 = driver.find_element(By.XPATH, '//*[@id="ui-id-2"]/li/a[text() = "'+ str(SEDOlog) +'"]')
    user1.click()
    psw.send_keys(SEDOpass)
    login.click()
    time.sleep(2)
    last_msg_id = asyncio.run(info_msg(UserID, 'Вход в СЭДО успешно выполнен', last_msg_id))
    # bot.send_message(UserID,"Вход в СЭДО успешно выполнен")

    driver.find_element(By.XPATH, '//*[@id="s-menu-stat"]').click()
    driver.find_element(By.XPATH, '//*[@id="s-menu-stat"]/div/a[1]').click()
    # Выбор года документов
    time.sleep(1)
    driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/a[14]').click()

    try:
        WebDriverWait(driver,600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="start_date"]'))).click()
    except:
        print("Timed out waiting for page to load")
    driver.find_element(By.XPATH, '//*[@id="start_date"]').click()
    driver.find_element(By.XPATH, '//*[@id="start_date"]').send_keys(Keys.CONTROL, 'a')
    driver.find_element(By.XPATH, '//*[@id="start_date"]').send_keys(Keys.BACKSPACE)
    driver.find_element(By.XPATH, '//*[@id="start_date"]').send_keys(str(date22))

    driver.find_element(By.XPATH, '//*[@id="end_date"]').click()
    driver.find_element(By.XPATH, '//*[@id="end_date"]').send_keys(Keys.CONTROL, 'a')
    driver.find_element(By.XPATH, '//*[@id="end_date"]').send_keys(Keys.BACKSPACE)
    driver.find_element(By.XPATH, '//*[@id="end_date"]').send_keys(str(date11))

    driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/form/div[1]/div[1]/a').click()

    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').click()
    Usr = User.split(sep='.')
    USR = str(Usr[0]).split()
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').send_keys(USR[0]+' ')
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').send_keys(USR[1]+'.')
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').send_keys(Usr[1]+'.')
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/ul/li[1]/div').click()

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="immediately"]').click()
    try:
        WebDriverWait(driver,600).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr')))
    except:
        print("Timed out waiting for page to load")
    time.sleep(1)
    maxrow = len(driver.find_elements(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr'))
    z=0
    df = pd.DataFrame()
    for i in range (1, maxrow):
        print(i)
        try:
            href = driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr['+ str(i) +']/td[5]/a').get_attribute('href')
            print (href)
            href = str(href)
            #driver.execute_script("window.open("+ href +")")
            driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr['+ str(i) +']/td[5]/a').send_keys(Keys.CONTROL, Keys.RETURN)
            b = driver.current_window_handle
            print(b)
            s = driver.window_handles
            print(s)
            for j in range (len(s)):
                print (j)
                if str(s[j]) != str(b):
                    z=z+1
                    a = str(s[j])
                    driver.switch_to.window(a)
                    try:
                        WebDriverWait(driver,600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="nevyp_nast"]')))
                    except:
                        print("Timed out waiting for page to load")
                    driver.find_element(By.XPATH, '//*[@id="nevyp_nast"]').click()
                    time.sleep(5)
                    driver.close()
                    driver.switch_to.window(b)
                    macroname = 'Красный контроль.xlsm'
                    polni = 'red control ' + str(z)
                    ########################################################################################################################################################################
                    path = 'C:\\Users\\'+PCuser+'\\Downloads'
                    files = os.listdir(path)
                    files = [os.path.join(path, file) for file in files]
                    files = [file for file in files if os.path.isfile(file)]
                    fp = max(files, key=os.path.getctime)
                    (df_in, ) = pd.read_html(fp)
                    try:
                        os.remove(fp)
                    except:
                        pass
                    print (df_in)
                    df_in = df_in.dropna()
                    df = pd.concat([df, df_in], axis = 0)
                    print (fp)
    #                move_xl(polni, foldername, macroname, PCuser)
                    last_msg_id = asyncio.run(info_msg(UserID, 'Красный контроль часть ' + str(z), last_msg_id))
                    # bot.send_message(UserID, 'Красный контроль часть ' + str(z))
        except:
            print('OK')

    #techpath2 = os.path.join(workdir, 'income data', 'technical_data', 'red control')

    #shutil.copy(workdir + '\\income data\\technical_data\\red control.xlsx', techpath2)





    #try:
    #    os.remove(techpath2 + r'\красный контроль.xlsx')
    #except:
    #    print(1)
    #try:
    #    os.remove(techpath2 + r'\Красный контроль 2.xlsm')
    #except:
    #    print(1)
    #r = os.listdir(techpath2)
    #path = techpath2 + '\\'
    #print(r)
    #df = pd.read_excel(str(path)+str(r[0]), sheet_name = 'Восстановл_Лист1')
    #
    #print (len(r))
    #for i in range (1, (len(r))):
    #    print(i)
    #    df1 = pd.read_excel(str(path)+str(r[i]), sheet_name = 'Восстановл_Лист1')
    #    df = pd.concat([df,df1])
        #print (df)
    print (df.info)

    techpath2 = os.path.join(workdir, 'income data', 'technical_data', 'red control')
    #df = df.set_index(['№ п/п'])
    df = df.reset_index(drop=True)
    #df = df.set_index(['№ п/п'])
    df['№ п/п'] = np.arange(1, 1 + len(df))

    df_copy = df.copy()
    mylist = df_copy['С установленным сроком исполнения'].to_list()
    mylist = [int(i.split()[2]) for i in mylist]

    df_copy = df_copy.assign(F = mylist)
    df_new = df_copy[df_copy['F'] > 59]
    redlist = df_new['№ п/п'].to_list()
    #df['Дата поручения'] = df['Дата поручения'].dt.strftime('%d.%m.%Y')
    df.to_excel(techpath2 + r'\красный контроль.xlsx', index = False, sheet_name = 'sheet1', engine = 'openpyxl')



    print (df_new)
    print(redlist)
    wb = load_workbook(techpath2 + r'\красный контроль.xlsx')
    ws = wb.active
    redFill = PatternFill(start_color='FFA07A',
                       end_color='FFA07A',
                       fill_type='solid')
    for i in redlist:
        cell = 'B' + str(i+1)
        print(cell)
        ws.cell(row = i+1, column = 2).fill = redFill
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['G'].width = 43
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 37
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 15
    wb.save(techpath2 + r'\красный контроль.xlsx')
    wb.close()
    shutil.copy(workdir + '\\income data\\technical_data\\Красный контроль 2.xlsm', techpath2)
    xl=win32.Dispatch("Excel.Application")
    macroname = 'Красный контроль 2.xlsm'
    xl.Workbooks.Open(os.path.abspath(techpath2 +'\\'+ macroname), ReadOnly=1)
    xl.Application.Quit() # Comment this out if your excel script closes
    del xl
    dater = now.date()
    try:
        shutil.copy(techpath2 + r'\красный контроль.xlsx', foldername + 'Красный контроль ' + str(dater) + '.xlsx')
    except:
        os.remove(foldername + 'Красный контроль ' + str(dater) + '.xlsx')

        shutil.copy(techpath2 + r'\красный контроль.xlsx', foldername + 'Красный контроль ' + str(dater) + '.xlsx')
    last_msg_id = asyncio.run(info_msg(UserID, 'Красный контроль завершен', last_msg_id))
    # bot.send_message(UserID, 'красный контроль успешно выполнен')
    redmail = str(len(df))
    redredmail = str(len(df_new))
    # try:
    #     bot.send_message(UserID, 'Просроченных писем: ' + str(len(df_new)))
    #     bot.send_message(UserID, 'Всего писем: ' + str(len(df)))
    # except:
    #     pass
    driver.close()


    now = datetime.today()

    end_date = now + timedelta(days = 30)
    date1 = str(end_date.date())
    date1 = date1.split(sep='-')
    date2 = str(now.date())
    date2 = date2.split(sep='-')
    date22 = date2[2]+'.'+date2[1]+'.'+date2[0]
    date11 = date1[2]+'.'+date1[1]+'.'+date1[0]
    date222 = date2[1]+'.'+date2[2]
    date111 = date1[1]+'.'+date1[2]
    startdateM = now
    date3 = str(startdateM.date())
    date3 = date3.split(sep='-')
    date33 = date3[2]+'.'+date3[1]+'.'+date3[0]
    date333 = date3[1]+'.'+date3[2]
    print (date1)
    print (date3)
    finpath = 'C:\\Users\\' + PCuser + '\\Desktop\\'

    if os.path.exists(finpath + 'Контроль писем общий') == False:
        os.mkdir(finpath + 'Контроль писем общий')
        #os.rename()
    finpath = 'C:\\Users\\' + PCuser + '\\Desktop\\Контроль писем общий\\'
    foldername = finpath + "Контроль " + str(date333) + '-' + str(date111) + '\\'
    if os.path.exists(foldername) == False:
        os.mkdir(foldername)
        # bot.send_message(UserID, 'Создал папку для предупредительного контроля')
        # bot.send_message(UserID, foldername)



    path = os.getcwd()
    r = os.listdir(os.path.join(workdir, 'income data', 'technical_data', 'pred control'))
    for i in range (len(r)):
        try:
            os.remove(str(os.path.join(workdir, 'income data', 'technical_data', 'pred control'))+'\\'+str(r[i]))
        except:
            pass
    year1 = 2020
    year2 = now.year
    chromedriver = str(path) + '/chromedriver.exe'
    driver = webdriver.Chrome()
    driver.get('https://mosedo.mos.ru/auth.php?uri=%2F')
    time.sleep(3)
    organization = driver.find_element(By.XPATH, '//*[@id="organizations"]')
    user = driver.find_element(By.XPATH, '//*[@id="logins"]')
    psw = driver.find_element(By.XPATH, '//*[@id="password_input"]')
    login = driver.find_element(By.XPATH, '//*[@id="login_form"]/table/tbody/tr[11]/td[4]/input[1]')
    organization.send_keys('Департамент городского имущества города Москвы')
    time.sleep(2)
    organization1 = driver.find_element(By.XPATH, '//*[@id="ui-id-1"]/li/a[text() = "Департамент городского имущества города Москвы"]')
    organization1.click()
    user.send_keys(SEDOlog)
    time.sleep(3)
    user1 = driver.find_element(By.XPATH, '//*[@id="ui-id-2"]/li/a[text() = "'+ str(SEDOlog) +'"]')
    user1.click()
    psw.send_keys(SEDOpass)
    login.click()
    time.sleep(2)
    last_msg_id = asyncio.run(info_msg(UserID, 'Вход в СЭДО успешно выполнен', last_msg_id))
    # bot.send_message(UserID,"Вход в СЭДО успешно выполнен")

    driver.find_element(By.XPATH, '//*[@id="s-menu-stat"]').click()
    driver.find_element(By.XPATH, '//*[@id="s-menu-stat"]/div/a[1]').click()
    # Выбор года документов
    time.sleep(1)
    driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/a[14]').click()

    try:
        WebDriverWait(driver,600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="start_date"]'))).click()
    except:
        print("Timed out waiting for page to load")
    driver.find_element(By.XPATH, '//*[@id="start_date"]').click()
    driver.find_element(By.XPATH, '//*[@id="start_date"]').send_keys(Keys.CONTROL, 'a')
    driver.find_element(By.XPATH, '//*[@id="start_date"]').send_keys(Keys.BACKSPACE)
    driver.find_element(By.XPATH, '//*[@id="start_date"]').send_keys(str(date22))

    driver.find_element(By.XPATH, '//*[@id="end_date"]').click()
    driver.find_element(By.XPATH, '//*[@id="end_date"]').send_keys(Keys.CONTROL, 'a')
    driver.find_element(By.XPATH, '//*[@id="end_date"]').send_keys(Keys.BACKSPACE)
    driver.find_element(By.XPATH, '//*[@id="end_date"]').send_keys(str(date11))

    driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/form/div[1]/div[1]/a').click()

    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').click()
    Usr = User.split(sep='.')
    USR = str(Usr[0]).split()
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').send_keys(USR[0]+' ')
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').send_keys(USR[1]+'.')
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/div/div[2]/input').send_keys(Usr[1]+'.')
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="select-user-adapter-author"]/div/div/div[1]/ul/li[1]/div').click()
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)
    driver.find_element(By.XPATH, '//*[@id="immediately"]').click()
    try:
        WebDriverWait(driver,600).until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr')))
    except:
        print("Timed out waiting for page to load")
    time.sleep(1)
    maxrow = len(driver.find_elements(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr'))
    z=0
    df = pd.DataFrame()
    for i in range (1, maxrow):
        print(i)
        try:
            href = driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr['+ str(i) +']/td[6]/a').get_attribute('href')
            print (href)
            href = str(href)
            #driver.execute_script("window.open("+ href +")")
            driver.find_element(By.XPATH, '/html/body/div[5]/div[3]/div[2]/div/div[3]/div/div/table/tbody/tr['+ str(i) +']/td[6]/a').send_keys(Keys.CONTROL, Keys.RETURN)
            b = driver.current_window_handle
            print(b)
            s = driver.window_handles
            print(s)
            for j in range (len(s)):
                print (j)
                if str(s[j]) != str(b):
                    z=z+1
                    a = str(s[j])
                    driver.switch_to.window(a)
                    try:

                        WebDriverWait(driver,600).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="nevyp_nenast"]')))
                    except:
                        print("Timed out waiting for page to load")
                    driver.find_element(By.XPATH, '//*[@id="nevyp_nenast"]').click()
                    time.sleep(5)
                    driver.close()
                    driver.switch_to.window(b)
                    macroname = 'Красный контроль.xlsm'
                    polni = 'pred control ' + str(z)
                    path = 'C:\\Users\\'+PCuser+'\\Downloads'
                    files = os.listdir(path)
                    files = [os.path.join(path, file) for file in files]
                    files = [file for file in files if os.path.isfile(file)]
                    fp = max(files, key=os.path.getctime)
                    (df_in, ) = pd.read_html(fp)
                    df_in = df_in.dropna()
                    df = pd.concat([df, df_in], axis=0)
                    try:
                        os.remove(fp)
                    except:
                        pass
                    last_msg_id = asyncio.run(info_msg(UserID, 'Предупредительный контроль часть ' + str(z), last_msg_id))
                    # bot.send_message(UserID, 'Предупредительный контроль часть ' + str(z))
    #                move_xl(polni, foldername, macroname, PCuser)


        except:
            print('OK')
    techpath2 = os.path.join(workdir, 'income data', 'technical_data', 'pred control')

    try:
        os.remove(techpath2 + r'\предупредительный контроль.xlsx')
    except:
        print(1)
    try:
        os.remove(techpath2 + r'\Предупредительный контроль 2.xlsm')
    except:
        print(1)

    #shutil.copy(workdir + '\\income data\\technical_data\\pred control.xlsx', techpath2)
    #r = os.listdir(techpath2)
    #df = pd.read_excel(str(techpath2)+'\\'+str(r[0]), sheet_name = 'Восстановл_Лист1')
    #
    #for i in range (1, (len(r))):
    #    df1 = pd.read_excel(str(techpath2)+'\\'+str(r[i]), sheet_name = 'Восстановл_Лист1')
    #    df = pd.concat([df,df1])
    #    print (df)

    #df = df.set_index(['№ п/п'])
    df = df.reset_index(drop=True)

    #df = df.set_index(['№ п/п'])
    df['№ п/п'] = np.arange(1, 1 + len(df))
    namexl = 'предупредительный контроль.xlsx'
    #df['Дата поручения'] = df['Дата поручения'].dt.strftime('%d.%m.%Y')
    df.to_excel(os.path.join(techpath2, namexl), index = False, engine = 'openpyxl')



    wb = load_workbook(techpath2 + r'\предупредительный контроль.xlsx')
    ws = wb.active
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['C'].width = 17
    ws.column_dimensions['D'].width = 17
    ws.column_dimensions['E'].width = 17
    ws.column_dimensions['F'].width = 17
    ws.column_dimensions['G'].width = 43
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 37
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 15
    wb.save(techpath2 + r'\предупредительный контроль.xlsx')
    wb.close()
    shutil.copy(workdir + '\\income data\\technical_data\\Предупредительный контроль 2.xlsm', techpath2)
    xl=win32.Dispatch("Excel.Application")
    macroname = 'Предупредительный контроль 2.xlsm'
    xl.Workbooks.Open(os.path.abspath(techpath2 +'\\'+ macroname), ReadOnly=1)
    xl.Application.Quit() # Comment this out if your excel script closes
    del xl
    dater = now.date()


    try:
        shutil.copy(techpath2 + r'\предупредительный контроль.xlsx', foldername + 'Предупредительный контроль ' + str(dater) + '.xlsx')
    except:
        os.remove(foldername + 'Предупредительный контроль ' + str(dater) + '.xlsx')
        shutil.copy(techpath2 + r'\предупредительный контроль.xlsx', foldername + 'Предупредительный контроль ' + str(dater) + '.xlsx')
        last_msg_id = asyncio.run(info_msg(UserID, 'Предупредительный контроль завершен', last_msg_id))
    # bot.send_message(UserID, 'Предупредительный контроль успешно выполнен')
    print('FUCK YEAH')
    driver.close()

    ####################ОТПРАВКА ПИСЬМА

    login = r"HQ\ArsenevVD"
    password = "Vitosik02010201!"
    server = "owa.mos.ru"
    port = 587
    recipients = ['GibadulinMM@mos.ru', 'LikhachIA@mos.ru', 'MusienkoOA@mos.ru', 'SiluyanovaYP@mos.ru']
    #recipients = ['ArsenevVD@mos.ru']
    #cc = ['FatykhovaLM@mos.ru', 'SafinRR@mos.ru']
    cc = ['ArsenevVD@mos.ru', 'ArtemovaOI@mos.ru', 'GabitovDS@mos.ru', 'DmitrievaKI@mos.ru',
          'ZavyalovaIV@mos.ru', 'IshinkinaEA@mos.ru', 'LavrentievaLV@mos.ru', 'NesterenkoAI1@mos.ru',
          'NikolenkoDA@mos.ru', 'OstapenkoAV2@mos.ru', 'SukhanovaYR@mos.ru', 'GoncharovaIA2@mos.ru',
          'InshakovaMN@mos.ru', 'BorisovaIN1@mos.ru']

    mail_date = dater.strftime('%d.%m.%Y')
    predpath = foldername + 'Предупредительный контроль ' + str(dater) + '.xlsx'
    redpath = foldername + 'Красный контроль ' + str(dater) + '.xlsx'
    subject = 'Красный и предупредительный контроли ' + mail_date
    body = 'Доброе утро! \n \nПо состоянию на ' + mail_date + ' за Управлением ведения жилищного учета числятся просроченные обращения в количестве: ' + redmail + ' шт. \nИз них количество обращений, срок исполнения которых истек более 2 месяцев назад составляет: ' + redredmail + ' шт. \n \nВо избежание проблем в части нарушения исполнительской дисциплины в Управлении, прошу обратить особое внимание на данный контроль и усилить меры по исполнению и закрытию вышеуказанных обращений. \n \nВ случае наличия объективных причин для неисполнения обращения в срок, необходимо продлить контрольный срок по согласованию с начальником Управления.'
    time.sleep(3)
    last_msg_id = asyncio.run(info_msg(UserID, 'Предупредительный контроль завершен', last_msg_id))
    att_paths = [redpath, predpath]
    time.sleep(3)
    last_msg_id = asyncio.run(info_msg(UserID, f'Направьте письмо с указанным текстом адресатам. Вложения придут в следующих сообщениях. Текст письма:\n{body}', last_msg_id))
    for i in range(len(att_paths)):
        print(att_paths[i])
        file = FSInputFile(att_paths[i])
        asyncio.run(bot.send_document(UserID, file, caption=f'Красный и предупредительный контроли файл {str(i+1)} из {len(att_paths)}'))
        try:
            asyncio.run(bot.session.close())
        except:
            pass
    k = 0
    ind = 0

    # SMTPmail.send_email(login, password, server, port, recipients, cc, subject, body, attachment_paths)
    # bot.send_message(UserID, 'Письмо направлено')
    #send_email(recipients, subject, body, attachment_paths, cc )
    # while k < 3:
    #     try:
    #         SMTPmail.send_email(login, password, server, port, recipients, cc, subject, body, attachment_paths)
    #         SMTPmail.send_email()
    #         k = 3
    #         bot.send_message(UserID, 'Письмо направлено')
    #     except:
    #         time.sleep(70)
    #         k = k+1
    #         ind = ind+1
    # if ind == 3:
    #     bot.send_message(UserID, 'Не удалось направить письмо')





