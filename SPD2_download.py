# -*- coding: utf-8 -*-
"""
Created on Mon Jun  5 16:05:57 2023

@author: ArsenevVD
"""



import psycopg2
from aiogram import Bot
from aiogram.enums import ParseMode
from openpyxl import Workbook
import pandas as pd
from datetime import datetime, timedelta
import pytz
import sys
import time
import json
import os
import SMTPmail
import config
# from handlers import bot
from aiogram.types import Message, FSInputFile
import asyncio
import multiprocessing as mp
from PyPDF2 import PdfReader
import requests
import config
bot = Bot(token=config.BOT_TOKEN, parse_mode=ParseMode.HTML)

def get_text(guid):

    if guid != '':

        text = ''

        try:

            reader = PdfReader(f'C:\\Users\\ArsenevVD\\Desktop\\control_mail_2.1\\DSA-main\\SPD2_docs\\pdf\\{guid}.pdf')
            number_of_pages = len(reader.pages)
            for i in range(number_of_pages):
                page = reader.pages[i]
                text = text + page.extract_text()
            print(guid)

        except:

            res = requests.get(f'https://dgi-spd.mos.ru/api/files/download/{guid}')
            with open(f'C:\\Users\\ArsenevVD\\Desktop\\control_mail_2.1\\DSA-main\\SPD2_docs\\pdf\\{guid}.pdf', 'wb') as f:
                f.write(res.content)
            reader = PdfReader(f'C:\\Users\\ArsenevVD\\Desktop\\control_mail_2.1\\DSA-main\\SPD2_docs\\pdf\\{guid}.pdf')

            number_of_pages = len(reader.pages)
            for i in range(number_of_pages):
                page = reader.pages[i]
                text = text + page.extract_text()
            print(guid)

        return text

    else:

        print('EMPTY!')
        return ''


def tz_change(column, df_final):
    try:
        #        df_final[column] = pd.to_datetime(df_final[column], format = '%Y-%m-%d %H:%M:%S')
        #        df_final[column] = df_final[column].apply(lambda x: x + timedelta(hours = 3))
        df_final[column] = pd.to_datetime(df_final[column], format='%Y-%m-%dT%H:%M:%S')
        df_final[column] = df_final[column].dt.tz_convert('Europe/Moscow')
        df_final[column] = df_final[column].dt.tz_localize(None)

    except:
        #        df_final[column] = pd.to_datetime(df_final[column], format = '%d.%m.%Y %H:%M:%S')
        #        df_final[column] = df_final[column].apply(lambda x: x + timedelta(hours = 3))
        df_final[column] = pd.to_datetime(df_final[column], format='%Y-%m-%dT%H:%M:%S')
        my_timezone = pytz.timezone('Europe/Moscow')
        df_final[column] = df_final[column].dt.tz_convert(my_timezone)
        df_final[column] = df_final[column].dt.tz_localize(None)
    return df_final


def list_otkaz(data_ENO):
    database = 'camunda'
    database2 = 'spd'

    connection = psycopg2.connect(
        host=config.hostname,
        port=config.port,
        database=database,
        user=config.username,
        password=config.password
    )

    connection2 = psycopg2.connect(
        host=config.hostname,
        port=config.port,
        database=database2,
        user=config.username,
        password=config.password
    )

    cursor2 = connection2.cursor()

    # Получите список всех таблиц в базе данных
    cursor2.execute("""
                    SELECT process_instance_key, start_date, subprocess_info_id FROM spd_process_task
                    WHERE process_instance_key IN %s AND task_name LIKE '%%Подготовить%%' AND subprocess_info_id IS NOT NULL
                    ORDER BY start_date DESC;
                    """, (data_ENO,))

    cursor = connection.cursor()

    rows2 = cursor2.fetchall()
    workbook = Workbook()
    sheet = workbook.active
    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor2.description]
    sheet.append(column_names)
    

    # Запишите данные
    for row in rows2:
        sheet.append(row)
    data = sheet.values
    columns = next(data)[0:]
    df_1 = pd.DataFrame(data, columns=columns)
    cursor2.close()
    connection2.close()
    print(df_1)
    data_camunda = tuple(df_1['subprocess_info_id'].to_list())

    # print(f"Data from table:")
    # for row in rows:
    #     print(row[1].tobytes().decode('utf-8'))

    cursor.execute("""
                    WITH ids AS 
                     (
                     SELECT proc_inst_id_, bytearray_id_
                     FROM act_hi_varinst
                     WHERE proc_inst_id_ IN %s AND name_ = 'flowCurrentResult'
                     )

                    SELECT * FROM ids
                    LEFT JOIN
                    (SELECT id_ as bytearray_id_, bytes_
                    FROM act_ge_bytearray
                    WHERE id_ IN(SELECT bytearray_id_ FROM ids)) AS idss
                    ON ids.bytearray_id_ = idss.bytearray_id_;
                    """, (data_camunda,))
    rows = cursor.fetchall()
    print(f"Data from table:")
    workbook = Workbook()
    sheet = workbook.active
    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    # for row in rows:
    #     print(row)
    #     row = list(row)
    #     row[3] = row[3].tobytes().decode('utf-8')
    #     sheet.append(row)

    for i in range(len(rows)):
        print(rows[i])
        rows[i] = list(rows[i])
        rows[i][3] = rows[i][3].tobytes().decode('utf-8')
        sheet.append(rows[i])
    data = sheet.values
    columns = next(data)[0:]

    df_2 = pd.DataFrame(rows, columns=columns)
    print(df_2)
    cursor.close()
    connection.close()
    df_fin = df_1.merge(df_2, left_on='subprocess_info_id', right_on='proc_inst_id_', how='left')

    df_fin['bytes_'] = df_fin['bytes_'].fillna('{"flowReason":"0"}')
    # df_fin['bytes_'] = df_fin['bytes_'].apply(lambda x: x.tobytes().decode('utf-8'))
    # json_list = df_fin['bytes_'].to_list()
    # for i in range(len(json_list)):
    #     try:
    #         json_list[i] = json.loads(json_list[i])['flowReason']
    #         print(json_list[i])
    #     except:
    #         json_list[i] = json_list[i]
    # df_fin['bytes_'] = json_list
    df_fin = tz_change('start_date', df_fin)
    result = df_fin.join(
        pd.json_normalize(
            df_fin.pop('bytes_').apply(
                str.replace, args=(',}', '}')
            ).apply(
                json.loads
            )
        )
    )
    print(result.columns)
    list_columns = result.columns
    df_col = pd.DataFrame(list_columns)
    needed_cols = ['start_date', 'flowConstructor.reason', 'flowConstructorCode',
                   'flowConstructor.block_3', 'flowConstructor.comment', 'flowConstructor.comment_1',
                   'flowConstructor.comment_11', 'flowConstructor.comment_12', 'flowConstructor.comment_13',
                   'flowConstructor.comment_17', 'flowConstructor.comment_2', 'flowConstructor.comment_3',
                   'flowConstructor.comment_33', 'flowConstructor.comment_4', 'flowConstructor.comment_5',
                   'flowConstructor.comment_6', 'flowConstructor.comment_end', 'flowConstructor.documents',
                   'flowConstructor.perechen_doc', 'flowConstructor.preambula', 'flowConstructor.preambulaText',
                   'flowConstructor.proizvol',
                   'virtualDocuments.main.Shablon_o_vnesenii_izmeneni_snyatii_itog.docx.data.otherList',
                   'virtualTempDocuments.main']
    cols = list(set(list_columns) & set(needed_cols))
    cols.insert(0, 'process_instance_key')
    df_col.to_excel(r'C:\Users\ArsenevVD\Downloads\columns.xlsx')
    result = result[cols]
    result = result[['process_instance_key', 'start_date', 'flowConstructor.reason',
                     'flowConstructorCode', 'virtualTempDocuments.main']]
    empty = [{'guid':''}]
    result['virtualTempDocuments.main'] = result['virtualTempDocuments.main'].fillna('empty_value')
    result['virtualTempDocuments.main'] = result['virtualTempDocuments.main'].apply(lambda x: empty if x == 'empty_value' else x)
    result['virtualTempDocuments.main'] = result['virtualTempDocuments.main'].apply(lambda x: x[0])
    result['virtualTempDocuments.main'] = result['virtualTempDocuments.main'].apply(lambda x: x['guid'])
    n_proc = 35
    numbers = result['virtualTempDocuments.main'].to_list()
    pool = mp.Pool(processes=n_proc)
    print(pool.map(get_text, numbers))
    pool.close()
    pool.join()
    result['virtualTempDocuments.main'] = result['virtualTempDocuments.main'].apply(lambda x: get_text(x))


    # df_fin['bytes_'] = df_fin['bytes_'].apply(lambda x: x.tobytes().decode('utf-8'))
    return result


def email_send(email, attach):
    login_mail = config.EMAIL_LOG
    password_mail = config.EMAIL_PASS
    server = "owa.mos.ru"
    port = 587
    recipients = [str(email)]
    cc = ['ArsenevVD@mos.ru']
    mail_date = datetime.strftime(datetime.today(), '%d.%m.%Y')
    subject = 'Выгрузка СПД-2 ' + mail_date
    body = 'Материалы, сформированные автоматически'
    SMTPmail.send_email(login_mail, password_mail, server, port, recipients, cc, subject, body, attach)
    return 'Письмо направлено на электронную почту'


def spd_2_download(email_, date_range_begin, date_range_end, sort_type, need_first_list, need_text, send_id):
    s = datetime.now()
    with open(os.getcwd() + '//settings.json') as f:
        settings = json.load(f)
        print(settings)
        #    time.sleep(20)
        PCuser = settings['PCuser']

    def tz_change(column, df_final):
        try:
            #        df_final[column] = pd.to_datetime(df_final[column], format = '%Y-%m-%d %H:%M:%S')
            #        df_final[column] = df_final[column].apply(lambda x: x + timedelta(hours = 3))
            df_final[column] = pd.to_datetime(df_final[column], format='%Y-%m-%dT%H:%M:%S')
            df_final[column] = df_final[column].dt.tz_convert('Europe/Moscow')
            df_final[column] = df_final[column].dt.tz_localize(None)

        except:
            #        df_final[column] = pd.to_datetime(df_final[column], format = '%d.%m.%Y %H:%M:%S')
            #        df_final[column] = df_final[column].apply(lambda x: x + timedelta(hours = 3))
            df_final[column] = pd.to_datetime(df_final[column], format='%Y-%m-%dT%H:%M:%S')
            my_timezone = pytz.timezone('Europe/Moscow')
            df_final[column] = df_final[column].dt.tz_convert(my_timezone)
            df_final[column] = df_final[column].dt.tz_localize(None)
        return df_final

    def convert_date(sort_type, date_range_begin, date_range_end):
        if sort_type == 0:
            a = datetime.strftime(datetime.strptime(date_range_begin, '%d.%m.%Y'), '%Y-%m-%d') + " 00:00:00+03"
            b = datetime.strftime(datetime.strptime(date_range_end, '%d.%m.%Y'), '%Y-%m-%d') + " 23:59:59+03"
        elif sort_type == 1:
            a = datetime.strftime(datetime.strptime(date_range_begin, '%d.%m.%Y'), '%Y-%m-%d')
            b = datetime.strftime(datetime.strptime(date_range_end, '%d.%m.%Y'), '%Y-%m-%d')
        return a, b

    # if date_range_end < date_range_begin:
    #     print("d1 > d2! Swapping!")
    #     date_range_begin, date_range_end = date_range_end, date_range_begin

    t1 = datetime.now()
    with open(os.getcwd() + '//settings.json') as f:
        settings = json.load(f)
        print(settings)
        PCuser = settings['PCuser']

    # Установите параметры подключения к базе данных

    database = 'spd_old_adapter'


    # Установите соединение с базой данных
    connection = psycopg2.connect(
        host=config.hostname,
        port=config.port,
        database=database,
        user=config.username,
        password=config.password
    )

    # Создайте курсор для выполнения SQL-запросов
    cursor = connection.cursor()

    # Получите список всех таблиц в базе данных
    cursor.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public'
        AND table_type = 'BASE TABLE'
    """)

    tables = cursor.fetchall()

    table_name = 'spd_act_documents'
    cursor.execute(f"SELECT * FROM {table_name}")  #### Расшифровки ГУ и источника поступления
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    for row in rows:
        sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_act_documents = pd.DataFrame(data, columns=columns)

    table_name = 'spd_task_adapter'

    cursor.execute(
        f"SELECT uuid, application_appdate, application_appname, application_parent_appdate, application_parent_appname FROM {table_name}")  #### Расшифровки ГУ и источника поступления
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    for row in rows:
        sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_task_adapter = pd.DataFrame(data, columns=columns)

    # Закройте курсор и соединение с базой данных
    cursor.close()
    connection.close()

    # Установите параметры подключения к базе данных

    database = 'camunda'


    # Установите соединение с базой данных
    connection = psycopg2.connect(
        host=config.hostname,
        port=config.port,
        database=database,
        user=config.username,
        password=config.password
    )

    # Создайте курсор для выполнения SQL-запросов
    cursor = connection.cursor()

    # Получите список всех таблиц в базе данных
    cursor.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public'
        AND table_type = 'BASE TABLE'
    """)

    tables = cursor.fetchall()

    table_name = 'spd_act_documents'
    cursor.execute(
        f"SELECT root_proc_inst_id_, name_, create_time_, text_ FROM act_hi_varinst WHERE name_ = 'rdNumber' ")  #### Расшифровки ГУ и источника поступления
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    for row in rows:
        sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_rds = pd.DataFrame(data, columns=columns)
    df_rds['create_time_'] = df_rds['create_time_'] + timedelta(hours=3)
    df_rds.drop(columns=['name_'], inplace=True)
    df_rds.sort_values(by='create_time_', ascending=False, inplace=True)
    df_rds.drop_duplicates(subset='root_proc_inst_id_', keep='first', inplace=True)

    cursor.close()
    connection.close()

    # Установите параметры подключения к базе данных

    database = 'spd'


    # Установите соединение с базой данных
    connection = psycopg2.connect(
        host=config.hostname,
        port=config.port,
        database=database,
        user=config.username,
        password=config.password
    )

    # Создайте курсор для выполнения SQL-запросов
    cursor = connection.cursor()

    # Получите список всех таблиц в базе данных
    cursor.execute("""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public'
        AND table_type = 'BASE TABLE'
    """)

    tables = cursor.fetchall()

    a, b = convert_date(sort_type, date_range_begin, date_range_end)

    table_name = 'spd_process_instance'
    print(a)
    print(b)
    data = (a, b,)

    if sort_type == 0:
        query = """SELECT business_key, process_instance_id, service_code, process_status, application_date, end_date, applicant_name, applicant_personal, due_date, due_days, real_due_date, targets, reject_accept_documents_due_date, reg_number, reg_date, spd_internal_number, deleter 
                   FROM spd_process_instance 
                   WHERE end_date >= %s AND end_date <= %s
                   ORDER BY end_date DESC"""
    elif sort_type == 1:
        query = """SELECT business_key, process_instance_id, service_code, process_status, application_date, end_date, applicant_name, applicant_personal, due_date, due_days, real_due_date, targets, reject_accept_documents_due_date, reg_number, reg_date, spd_internal_number, deleter 
                   FROM spd_process_instance
                   WHERE reg_date >= %s AND reg_date <= %s
                   ORDER BY reg_date DESC"""

    cursor.execute(query, data)  #### Расшифровки ГУ и источника поступления
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    for row in rows:
        sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_process = pd.DataFrame(data, columns=columns)
    print(df_process)
    df_process = df_process[df_process['deleter'].isnull()]
    col_list = ['application_date', 'end_date', 'due_date', 'real_due_date', 'reject_accept_documents_due_date']

    for item in col_list:
        df_process = tz_change(item, df_process)

    table_name = 'spd_address'

    cursor.execute(f"SELECT business_key, address, flat FROM {table_name}")  #### Расшифровки ГУ и источника поступления
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    for row in rows:
        sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_address = pd.DataFrame(data, columns=columns)

    table_name = 'spd_suspension'

    cursor.execute(
        f"SELECT business_key, start_date, end_date, real_end_date FROM {table_name}")  #### Расшифровки ГУ и источника поступления
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    # Запишите заголовки столбцов
    column_names = [desc[0] for desc in cursor.description]
    sheet.append(column_names)

    # Запишите данные
    for row in rows:
        sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_suspensions = pd.DataFrame(data, columns=columns)
    susp_dates = ['start_date', 'end_date', 'real_end_date']

    for item in susp_dates:
        df_suspensions = tz_change(item, df_suspensions)

    table_name = 'spd_status'
    list_otkaz_gu = ['Услуга оказана. Решение отрицательное', 'Услуга оказана Решение отрицательное',
                     'Отказ в предоставлении услуги']
    list_OPD = ['Отказ в приеме документов', 'Отказ в приеме запроса и документов ОИВ']
    status = ['Услуга оказана. Решение отрицательное', 'Услуга оказана Решение отрицательное',
              'Отказ в приеме запроса и документов ОИВ', 'Отказ в приеме документов', 'Отказ в предоставлении услуги']

    workbook = Workbook()
    sheet = workbook.active
    for i in range(len(status)):
        cursor.execute(f"SELECT business_key, name FROM {table_name} WHERE name = %s",
                       (status[i],))  #### Расшифровки ГУ и источника поступления
        rows = cursor.fetchall()

        # Запишите заголовки столбцов
        column_names = [desc[0] for desc in cursor.description]
        sheet.append(column_names)

        # Запишите данные
        for row in rows:
            sheet.append(row)

    data = sheet.values
    columns = next(data)[0:]
    df_status = pd.DataFrame(data, columns=columns)

    cursor.close()
    connection.close()

    df_act_documents = df_act_documents.sort_values(by=['doc_date'])
    df_rubr = pd.read_excel(
        os.path.join(os.getcwd(),'income data', 'proc_targetss.xlsx'))  #######################################################################################ИСПРАВИТЬ!

    df_final = df_act_documents.merge(df_task_adapter, left_on='task_uuid', right_on='uuid', how='left')
    del df_task_adapter
    df_final = df_process.merge(df_final, left_on='reg_number', right_on='application_parent_appname', how='left')
    del df_process
    df_final = df_final.merge(df_address, left_on='business_key', right_on='business_key', how='left')
    del df_address
    df_final['targets'] = df_final['targets'].apply(lambda x: str(x).strip('][').split(', ')[0])
    df_final['targets'] = df_final['targets'].apply(lambda x: int(x.split(',')[0]) if x != 'None' else 0)
    df_final = df_final.drop_duplicates(subset=['reg_number'], keep='first')
    df_rubr['targets_code'] = df_rubr['targets_code'].apply(lambda x: int(x))
    df_rubr['service_code'] = df_rubr['service_code'].apply(lambda x: int(x))

    df_final['service_code'] = df_final['service_code'].apply(lambda x: int(x))
    df_final = df_final.merge(df_rubr, left_on=['service_code', 'targets'], right_on=['service_code', 'targets_code'],
                              how='left')
    del df_rubr
    df_final = df_final.merge(df_status, left_on='business_key', right_on='business_key', how='left')
    del df_status

    ###################################################################################################################################### Правки от 15.11.2023
    df_final = df_final.merge(df_rds, left_on='process_instance_id', right_on='root_proc_inst_id_', how='left')
    df_final.drop(columns=['process_instance_id', 'root_proc_inst_id_'])
    ###################################################################################################################################### Правки от 15.11.2023

    df_final = df_final.merge(df_suspensions, left_on='business_key', right_on='business_key', how='left')
    del df_suspensions
    suspensions_count = df_final.groupby(['reg_number']).size().reset_index(name='Количество приостановок')
    df_final = pd.merge(df_final, suspensions_count, on=['reg_number'], how='left')
    df_final.loc[df_final['start_date'].isna(), 'Количество приостановок'] = 0
    df_final.info()

    df_final.rename(columns={'source_name': 'Источник поступления',
                             'reg_number': 'Номер заявки ДГИ',
                             'business_key': 'ЕНО',
                             'applicant_name': 'Заявитель',
                             'applicant_personal': 'Лицо',
                             'application_date': 'Дата поступления заявки',
                             'reg_date': 'Дата регистрации заявки',
                             'spdLink': 'Ссылка',
                             'process_status': 'Статус заявки',
                             'service_name': 'Первые "Statistics"[Госуслуга]',
                             'due_date': 'Дата контроля заявки',
                             'end_date_x': 'Дата принятия решения',
                             'start_date': 'Дата начала приостановки',
                             'end_date_y': 'Дата окончания приостановки (план)',
                             'real_end_date': 'Дата окончания приостановки (факт)',
                             'targets_y': 'Цель',
                             'doc_date': 'Дата документа',
                             'doc_number': 'Номер документа',
                             'doc_type': 'Тип документа',
                             'application_parent_appdate': 'Дата родительской заявки',
                             'application_parent_appname': 'Номер родительской заявки',
                             'address': 'Адрес',
                             'flat': 'Квартира номер',
                             'create_time_': 'Дата РД',
                             'text_': 'Номер РД'
                             }, inplace=True)

    df_final.drop(columns=['targets_x', 'uuid_x', 'task_uuid', 'uuid_y', 'targets_code',
                           'application_appdate', 'application_appname',
                           'service_code', 'spd_internal_number', 'reject_accept_documents_due_date', ], inplace=True)
    neworder = ['ЕНО', 'Источник поступления', 'Дата поступления заявки',
                'Номер заявки ДГИ', 'Дата регистрации заявки', 'Дата контроля заявки',
                'Дата принятия решения', 'Статус заявки',
                'Первые "Statistics"[Госуслуга]', 'Цель', 'Заявитель', 'Лицо', 'Адрес',
                'Квартира номер', 'Дата документа', 'Номер документа',
                'Тип документа', 'Дата родительской заявки', 'Номер родительской заявки', 'Количество приостановок',
                'Дата начала приостановки', 'Дата окончания приостановки (план)',
                'Дата окончания приостановки (факт)', 'due_days', 'real_due_date', 'name', 'Номер РД', 'Дата РД']
    df_final = df_final.reindex(columns=neworder)

    df_final.loc[df_final['Лицо'] == True, 'Лицо'] = "Физическое"
    df_final.loc[df_final['Статус заявки'] == "APPROVED", 'Статус заявки'] = "Положительное решение"
    df_final.loc[df_final['Статус заявки'] == "CANCELED", 'Статус заявки'] = "Отзыв заявителем"
    df_final.loc[df_final['Статус заявки'] == "IN_PROGRESS", 'Статус заявки'] = "В работе"
    df_final.loc[df_final['Статус заявки'] == "REGISTERED", 'Статус заявки'] = "Зарегистрировано"
    df_final.loc[df_final['Статус заявки'] == "REJECTED", 'Статус заявки'] = "Отказ в услуге / в приеме документов"
    df_final['name'].fillna('')
    df_final.loc[
        df_final['name'].str.contains('|'.join(list_OPD), na=False), 'Статус заявки'] = "Отказ в приеме документов"
    df_final.loc[df_final['name'].str.contains('|'.join(list_otkaz_gu),
                                               na=False), 'Статус заявки'] = "Отказ в предоставлении услуги"
    df_final.loc[df_final[
                     'Статус заявки'] == "Отказ в услуге / в приеме документов", 'Статус заявки'] = "Отказ в приеме документов"
    df_final.loc[df_final['Статус заявки'] == "SUSPENDED", 'Статус заявки'] = "Приостановлено"
    df_final.drop(columns='name', inplace=True)
    df_final.loc[df_final['Первые "Statistics"[Госуслуга]'].str.contains(
        'Заключение договоров социального найма жилых помещений жилищного фонда города Москвы, предоставленных по ордеру или на основании иных решений органов исполнительной власти города Москвы, заключение дополнительных соглашений к договорам безвозмездного пользования жилыми помещениями'), 'Первые "Statistics"[Госуслуга]'] = "Заключение договоров социального найма жилых помещений жилищного фонда города Москвы, предоставленных по ордеру или на основании иных решений органов исполнительной власти города Москвы, заключение дополнительных соглашений к договорам безвозмездного пользования жилыми помещениями специализированного жилищного фонда города Москвы, социального найма жилых помещений жилищного фонда города Москвы, найма жилых помещений жилищного фонда коммерческого использования города Москвы"
    date = datetime.strftime(datetime.now(), '%Y.%m.%d')
    date_time = datetime.strftime(datetime.now(), '%Y.%m.%d %H.%M')
    df = df_final.drop_duplicates(subset=['Номер заявки ДГИ'], keep='first')
    if need_text == 1:
        data = tuple(df['ЕНО'].to_list())
        i = 0
        j = 0
        z = 0
        data1 = []
        result = pd.DataFrame()
        while i < len(data):
            print(i)
            if (z < 1000) and (i < len(data) - 1):
                data1.append(data[i])
                z = z + 1
            else:
                print(' zapusk ', i)
                df_out = list_otkaz(tuple(data1))
                result = pd.concat([result, df_out])
                # df_out.to_excel(f'C:\\Users\\ArsenevVD\\Downloads\\{str(i)}.xlsx')
                data1 = []
                z = 0
            i = i + 1
        result = pd.merge(df, result, left_on='ЕНО', right_on='process_instance_key', how='left')

    with pd.ExcelWriter(os.path.join(os.getcwd(), 'export_data', 'dashboard', f'dashboard {date_time}.xlsx')) as writer:
        if need_first_list == 1:
            df_final.to_excel(writer, sheet_name='с приостановками')
        if need_text == 1:
            result.to_excel(writer, sheet_name='с текстом')
        df_final = df_final.drop_duplicates(subset=['Номер заявки ДГИ'], keep='first')
        df_final.to_excel(writer, sheet_name='без дубликатов')

    att = os.path.join(os.getcwd(), 'export_data', 'dashboard', f'dashboard {date_time}.xlsx')
    file = FSInputFile(att)
    asyncio.run(bot.send_document(send_id, file, caption=f'Выгрузка СПД 2 с {date_range_begin} по {date_range_end}'))
    try:
        asyncio.run(bot.session.close())
    except:
        pass
    # resp = email_send(email_, [att])
    # asyncio.run(bot.send_message(chat_id=send_id, text='Выгрузка СПД-2 направлена на почту'))
    # asyncio.run(bot.session.close())
    print(f'Execution took {datetime.now() - t1}')
    return 'Выгрузка готова!'


if __name__ == "__main__":
    s = datetime.now()
    spd_2_download('ArsenevVD@mos.ru', '19.04.2024', '19.04.2024', 0, 0, 0, 260399228)

